"""Write all tables to CSV and a formatted Excel workbook."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def save_csv_tables(tables: dict[str, pd.DataFrame], tables_dir: Path) -> None:
    tables_dir.mkdir(parents=True, exist_ok=True)
    for name, df in tables.items():
        df.to_csv(tables_dir / f"{name}.csv", index=False)


def _format_sheet(ws, title: str, description: str, ncols: int) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="2B6CB0")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ncols, 1))

    ws["A2"] = description
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(ncols, 1))

    header_fill = PatternFill("solid", fgColor="E2E8F0")
    for col in range(1, ncols + 1):
        c = ws.cell(row=4, column=col)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, horizontal="center")

    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        width = 12
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                if cell.value is not None:
                    width = max(width, min(len(str(cell.value)) + 2, 48))
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A4"


def export_excel(tables: dict[str, pd.DataFrame], excel_path: Path) -> Path:
    """Excel workbook with one sheet per table (see docs)."""
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    plan = (
        ("All Predictors OR", "combined_or", "Primary result: adjusted OR and p-value for every predictor"),
        ("Study Overview", "overview", "Design summary and sample description"),
        ("Descriptives", "descriptives", "Continuous variable summary statistics"),
        ("Prev by SES", "prev_ses", "Unadjusted prevalence by SES tertile"),
        ("Prev by HISB", "prev_hisb", "Unadjusted prevalence by HISB group"),
        ("Prev by SelfTreat", "prev_self_treat", "Unadjusted prevalence by Self_Treat level"),
        ("SelfTreat Effect", "self_treat_effect", "Crude vs adjusted Self_Treat OR"),
        ("Adjusted ORs Detail", "adjusted_or_long", "Full adjusted model with 95% CIs"),
        ("Model Fit OTC", "model_fit_otc", "Model comparison for OTC outcome"),
        ("Model Fit Herbal", "model_fit_herbal", "Model comparison for herbal outcome"),
        ("Spearman", "spearman", "Self_Treat correlation with each outcome"),
    )
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        for sheet, key, desc in plan:
            df = tables.get(key)
            if df is None or df.empty:
                df = pd.DataFrame({"note": ["No data"]})
            df.to_excel(writer, sheet_name=sheet[:31], index=False, startrow=3)
            ws = writer.sheets[sheet[:31]]
            _format_sheet(ws, sheet, desc, len(df.columns))
    return excel_path
